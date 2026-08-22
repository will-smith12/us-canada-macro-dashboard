#!/usr/bin/env python3
"""
CFIB Business Barometer -> economics indicators refresher.

What it does
------------
1. Downloads the monthly CFIB "MBB data" workbook directly (no web scraping).
   The page at cfib-fcei.ca/.../business-barometer is just a rendering of this
   file, so we go straight to the source.
2. Parses the wide "Data" sheet (indicators down rows, months across columns,
   2009-02 -> present) and reshapes it into:
       - a tidy/long table:  Date | Group | Series | Value
       - a wide matrix:      Date x Series   (Date down the rows)
3. Writes cfib_business_barometer.xlsx (Tidy + Wide sheets).

Usage
-----
    python3 cfib_refresh.py                # auto-find latest published month
    python3 cfib_refresh.py --month 2026-05
    python3 cfib_refresh.py --file CFIB_MBB-data-2026-05.xlsx   # parse a local file

URL pattern (for reference / Power Query):
    https://www.cfib-fcei.ca/hubfs/research/mbb/{YYYY}/{MonthName}{YYYY}/CFIB_MBB-data-donnes-{YYYY}-{MM}.xlsx
e.g. May 2026 -> .../mbb/2026/May2026/CFIB_MBB-data-donnes-2026-05.xlsx
"""
from __future__ import annotations
import argparse, datetime as dt, sys, urllib.request
from pathlib import Path
import openpyxl
import pandas as pd

MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]

PROVINCES = {"Newfoundland & Lab.","Prince Edward Is.","Nova Scotia","New Brunswick",
             "Quebec","Ontario","Manitoba","Saskatchewan","Alberta","British Columbia"}
NATIONAL = {"Canada - outlook on 12 months","Canada - outlook on 3 months",
            "GDP%ch (SAAR)","Number of responses","Margin of error"}
# union of the sector names used in the Long-term and Short-term index blocks
SECTORS = {"Agriculture","Natural resources*","Construction","Manufacturing","Wholesale",
           "Retail","Transportation, warehousing","Info., arts, recreation",
           "Insurance, real estate, finance","Finance, insurance, real estate",
           "Prof., business services","Health & education serv.","Hospitality",
           "Personal services","Personal and other services"}


def build_url(year: int, month: int) -> str:
    return (f"https://www.cfib-fcei.ca/hubfs/research/mbb/{year}/"
            f"{MONTHS[month-1]}{year}/CFIB_MBB-data-donnes-{year}-{month:02d}.xlsx")


def download_latest(anchor: dt.date, lookback: int = 4) -> Path:
    """Try anchor month, then walk backwards (report lands ~3rd week of month)."""
    y, m = anchor.year, anchor.month
    for _ in range(lookback):
        url = build_url(y, m)
        out = Path(f"CFIB_MBB-data-{y}-{m:02d}.xlsx")
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as r:
                data = r.read()
            if data[:2] == b"PK":                       # valid xlsx (zip) magic
                out.write_bytes(data)
                print(f"  downloaded {url}")
                return out
        except Exception as e:
            print(f"  miss {y}-{m:02d}: {e}")
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    sys.exit("Could not find a published file in the lookback window.")


def _looks_like_data_sheet(ws) -> bool:
    """True if this sheet carries the monthly date header on row 3 from column 4."""
    try:
        header = [ws.cell(3, c).value for c in range(4, min(ws.max_column, 16) + 1)]
    except Exception:
        return False
    return sum(isinstance(v, dt.datetime) for v in header) >= 3


def _select_data_sheet(wb):
    """Find the CFIB data sheet.

    CFIB periodically renames the single data sheet ("Data" -> "Dtfile_new" ->
    "Datafile" -> ...), which used to break the hardcoded ["Data"] lookup. Pick
    the sheet by structure (most robust), then by name hint, then fall back to a
    lone sheet, so a future rename doesn't break the parse.
    """
    for ws in wb.worksheets:                      # 1) structural match (rename-proof)
        if _looks_like_data_sheet(ws):
            return ws
    for ws in wb.worksheets:                       # 2) name hint fallback
        norm = "".join(ch for ch in ws.title.lower() if ch.isalnum())
        if "data" in norm or "dtfile" in norm:
            return ws
    if len(wb.worksheets) == 1:                    # 3) single-sheet fallback
        return wb.worksheets[0]
    raise KeyError(f"CFIB data sheet not found (sheets={wb.sheetnames})")


def parse(path: Path) -> pd.DataFrame:
    ws = _select_data_sheet(openpyxl.load_workbook(path, data_only=True))
    date_cols = [(c, ws.cell(3, c).value) for c in range(4, ws.max_column + 1)
                 if isinstance(ws.cell(3, c).value, dt.datetime)]
    dates = [d.date() for _, d in date_cols]

    rows, horizon, question = [], None, None
    for r in range(4, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label or not str(label).strip():
            continue
        label = str(label).strip()
        vals = [ws.cell(r, c).value for c, _ in date_cols]
        has_data = any(isinstance(v, (int, float)) for v in vals)

        # the two index blocks are introduced by these header rows (no data)
        if label.startswith("Long-term index"):
            horizon = "Long-term"; continue
        if label.startswith("Short-term index"):
            horizon = "Short-term"; continue
        if not has_data:                       # any other no-data row = section/question
            question = label; continue

        if label in NATIONAL:
            group = "National headline"
        elif label in SECTORS:
            group = f"{horizon} index - Sectors"
        elif label in PROVINCES:
            group = f"{horizon} index - Provinces"
        else:                                  # survey % breakdowns, trade, etc.
            group = question or "Survey results"

        for d, v in zip(dates, vals):
            if isinstance(v, (int, float)):
                rows.append((d, group, label, float(v)))

    return pd.DataFrame(rows, columns=["Date", "Group", "Series", "Value"])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", help="YYYY-MM")
    ap.add_argument("--file", help="parse a local .xlsx instead of downloading")
    args = ap.parse_args()

    if args.file:
        path = Path(args.file)
    elif args.month:
        y, m = map(int, args.month.split("-"))
        path = download_latest(dt.date(y, m, 1), lookback=1)
    else:
        path = download_latest(dt.date.today().replace(day=1))

    tidy = parse(path)
    wide = (tidy.assign(col=tidy["Group"] + " | " + tidy["Series"])
                .pivot_table(index="Date", columns="col", values="Value")
                .sort_index())

    out = Path("cfib_business_barometer.xlsx")
    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        tidy.sort_values(["Date", "Group", "Series"]).to_excel(
            xl, sheet_name="Tidy", index=False)
        wide.to_excel(xl, sheet_name="Wide (Date x Series)")

    last = tidy["Date"].max()
    print(f"\nWrote {out}")
    print(f"  rows (tidy): {len(tidy):,}   series: {tidy['Series'].nunique()}   "
          f"months: {tidy['Date'].nunique()}   latest: {last}")
    head = tidy[(tidy.Date == last) & (tidy.Group == 'National headline')]
    for _, x in head.iterrows():
        print(f"    {x.Series:<32} {x.Value:,.2f}")


if __name__ == "__main__":
    main()
