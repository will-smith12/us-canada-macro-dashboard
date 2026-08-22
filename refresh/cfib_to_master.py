#!/usr/bin/env python3
"""
Wire the CFIB Business Barometer into master_economics_data.xlsx as dedicated sheets.

Adds (idempotently — re-running replaces them):
  - "CFIB Barometer"        wide, house style: Date down col A, two-row header
                            (row 3 = Group, row 4 = Series), data from row 5.
  - "CFIB Barometer (Tidy)" long format: Date | Group | Series | Value.

Full scope: every series with data in the source (national headline, long- and
short-term indices by sector and province, and all survey % breakdowns).

Existing sheets are left untouched; the two CFIB sheets are appended at the end.
Refuses to run if the workbook is open (lock file present).

Usage:
    python3 cfib_to_master.py --source CFIB_MBB-data-2026-05.xlsx
    python3 cfib_to_master.py            # downloads the latest month first
"""
from __future__ import annotations
import argparse, datetime as dt
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from cfib_refresh import parse, download_latest   # reuse the tested parser

TARGET = "master_economics_data.xlsx"
WIDE, TIDY = "CFIB Barometer", "CFIB Barometer (Tidy)"
DATEFMT = "yyyy-mm-dd"


def ordered_pairs(tidy):
    """(Group, Series) in first-appearance (source) order, not alphabetical."""
    seen, out = set(), []
    for g, s in zip(tidy["Group"], tidy["Series"]):
        if (g, s) not in seen:
            seen.add((g, s)); out.append((g, s))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source")
    ap.add_argument("--target", default=TARGET)
    args = ap.parse_args()

    tgt = Path(args.target)
    if (tgt.parent / f"~${tgt.name}").exists():
        raise SystemExit(f"ABORT: {tgt} appears to be open in Excel (lock file). Close it first.")

    src = Path(args.source) if args.source else download_latest(dt.date.today().replace(day=1))
    tidy = parse(src)
    pairs = ordered_pairs(tidy)
    dates = sorted(tidy["Date"].unique())
    src_month = max(dates)
    today = dt.date.today().isoformat()

    # value lookup: (group, series, date) -> value
    val = {(g, s, d): v for g, s, d, v in
           zip(tidy["Group"], tidy["Series"], tidy["Date"], tidy["Value"])}

    print(f"source={src.name}  series={len(pairs)}  months={len(dates)}  latest={src_month}")

    wb = openpyxl.load_workbook(tgt)               # default load preserves everything
    for name in (WIDE, TIDY):
        if name in wb.sheetnames:
            del wb[name]

    title_font = Font(bold=True, size=12)
    hdr_font = Font(bold=True)
    grp_fill = PatternFill("solid", fgColor="DDEBF7")
    sub_font = Font(italic=True, color="555555")
    src_note = (f"Source: Canadian Federation of Independent Business — Monthly Business "
                f"Barometer® ({src.name}). Monthly; dates = month start (CFIB convention). "
                f"Latest reading {src_month}. Wired in {today}.")

    # ---------- WIDE sheet (house style) ----------
    ws = wb.create_sheet(WIDE)
    ws["A1"] = "CFIB Business Barometer® — small-business confidence (Canada)"
    ws["A1"].font = title_font
    ws["A2"] = src_note
    ws["A2"].font = sub_font
    ws["A4"] = "Date"; ws["A4"].font = hdr_font
    for j, (g, s) in enumerate(pairs, start=2):
        cg = ws.cell(3, j, g); cg.font = hdr_font; cg.fill = grp_fill
        cg.alignment = Alignment(horizontal="center")
        cs = ws.cell(4, j, s); cs.font = hdr_font
        cs.alignment = Alignment(wrap_text=True, vertical="top")
    for i, d in enumerate(dates, start=5):
        dc = ws.cell(i, 1, d); dc.number_format = DATEFMT
        for j, (g, s) in enumerate(pairs, start=2):
            v = val.get((g, s, d))
            if v is not None:
                ws.cell(i, j, v)
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 12
    for j in range(2, len(pairs) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 13

    # ---------- TIDY sheet ----------
    wt = wb.create_sheet(TIDY)
    wt["A1"] = "CFIB Business Barometer® — tidy / long format"
    wt["A1"].font = title_font
    wt["A2"] = src_note; wt["A2"].font = sub_font
    for j, h in enumerate(["Date", "Group", "Series", "Value"], start=1):
        c = wt.cell(4, j, h); c.font = hdr_font; c.fill = grp_fill
    order = {gs: k for k, gs in enumerate(pairs)}
    recs = sorted(zip(tidy["Date"], tidy["Group"], tidy["Series"], tidy["Value"]),
                  key=lambda r: (order[(r[1], r[2])], r[0]))
    for i, (d, g, s, v) in enumerate(recs, start=5):
        dc = wt.cell(i, 1, d); dc.number_format = DATEFMT
        wt.cell(i, 2, g); wt.cell(i, 3, s); wt.cell(i, 4, v)
    wt.freeze_panes = "A5"
    for col, w in (("A", 12), ("B", 34), ("C", 30), ("D", 12)):
        wt.column_dimensions[col].width = w

    # place the two CFIB sheets right after "Business" if it exists, else leave at end
    cfib = [s for s in wb._sheets if s.title in (WIDE, TIDY)]
    rest = [s for s in wb._sheets if s.title not in (WIDE, TIDY)]
    idx = next((i for i, s in enumerate(rest) if s.title == "Business"), len(rest) - 1)
    wb._sheets = rest[: idx + 1] + cfib + rest[idx + 1:]

    wb.save(tgt)
    print(f"Saved {tgt}: +{WIDE!r} (wide, {len(pairs)} series x {len(dates)} months) "
          f"and +{TIDY!r} ({len(recs):,} rows).")


if __name__ == "__main__":
    main()
