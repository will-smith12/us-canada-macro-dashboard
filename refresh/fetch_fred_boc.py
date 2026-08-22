"""
Appends free FRED/ALFRED (US) and Bank of Canada Valet (CA) sheets to economics_data.xlsx.
No Trading Economics exports are used. Run AFTER fetch_economics.py (or standalone).

Sheets added/refreshed:
  - FRED Vintage - GDP / CPI / Payrolls : first release vs. latest revised (look-ahead-bias check)
  - US Financial Conditions             : rates, spreads, stress & inflation-expectation series
  - Canada (BoC)                        : FX, policy/prime rates, GoC benchmark bond yields
"""
import os
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FRED_KEY = os.environ.get("FRED_API_KEY", "")
WORKBOOK = "/Users/william.smith/Downloads/economics_data.xlsx"

# ---------- series definitions ----------
VINTAGE = [
    ("GDPC1",    "Real GDP — Vintage", "FRED Vintage - GDP",      "#,##0.0"),
    ("CPIAUCSL", "CPI (All Urban, SA)", "FRED Vintage - CPI",      "0.000"),
    ("PAYEMS",   "Nonfarm Payrolls",    "FRED Vintage - Payrolls", "#,##0"),
]

FINCOND = [
    ("NFCI",        "Chicago Fed NFCI (wk)",      "0.000"),
    ("STLFSI4",     "StL Fed Stress Index (wk)",  "0.000"),
    ("VIXCLS",      "VIX",                        "0.00"),
    ("DGS3MO",      "3M Treasury Yld",            "0.00"),
    ("DGS2",        "2Y Treasury Yld",            "0.00"),
    ("DGS10",       "10Y Treasury Yld",           "0.00"),
    ("T10Y2Y",      "10Y-2Y Spread",              "0.00"),
    ("T10Y3M",      "10Y-3M Spread",              "0.00"),
    ("SOFR",        "SOFR",                       "0.00"),
    ("EFFR",        "Eff. Fed Funds",             "0.00"),
    ("T10YIE",      "10Y Breakeven Infl.",        "0.00"),
    ("DFII10",      "10Y TIPS Real Yld",          "0.00"),
    ("BAMLH0A0HYM2","High-Yield OAS",             "0.00"),
    ("SAHMREALTIME","Sahm Rule (RT)",             "0.00"),
    ("DTWEXBGS",    "Broad USD Index",            "0.00"),
]

# Commodity / input-cost prices (FRED) — North American benchmarks for US & Canada
FRED_COMMODITY = [
    ("DCOILWTICO", "WTI Crude (USD/bbl)",           "0.00", "Daily"),
    ("DHHNGSP",    "Henry Hub Nat Gas (USD/MMBtu)", "0.00", "Daily"),
]

# Senior Loan Officer Opinion Survey (FRED) — US bank C&I lending standards
FRED_SLOOS = [
    ("DRTSCILM", "Net % Tightening C&I — Small Firms",      "0.0", "Quarterly"),
    ("DRTSCIS",  "Net % Tightening C&I — Large/Med Firms",  "0.0", "Quarterly"),
]

BOC = [
    ("FXUSDCAD",          "USD/CAD FX",                   "0.0000"),
    ("V39079",            "Target Overnight Rate (Policy)","0.00"),
    ("V80691311",         "Prime Rate",                   "0.00"),
    ("AVG.INTWO",         "CORRA Overnight",              "0.0000"),
    ("TB.CDN.90D.MID",    "GoC 3M T-bill Yield",          "0.00"),
    ("BD.CDN.2YR.DQ.YLD", "GoC 2Y Yield",                 "0.00"),
    ("BD.CDN.5YR.DQ.YLD", "GoC 5Y Yield",                 "0.00"),
    ("BD.CDN.10YR.DQ.YLD","GoC 10Y Yield",                "0.00"),
    ("BD.CDN.LONG.DQ.YLD","GoC Long Yield",               "0.00"),
]

NEW_SHEETS = ([v[2] for v in VINTAGE] +
              ["US Financial Conditions", "US SLOOS", "Commodity Prices", "Canada (BoC)"])

# ---------- API helpers ----------
def fred_obs(series_id, **params):
    p = {"series_id": series_id, "api_key": FRED_KEY, "file_type": "json"}
    p.update(params)
    r = requests.get("https://api.stlouisfed.org/fred/series/observations", params=p, timeout=60)
    r.raise_for_status()
    return r.json().get("observations", [])

def fred_earliest_vintage(series_id):
    r = requests.get("https://api.stlouisfed.org/fred/series/vintagedates",
                     params={"series_id": series_id, "api_key": FRED_KEY,
                             "file_type": "json", "limit": 1}, timeout=30)
    r.raise_for_status()
    vd = r.json().get("vintage_dates", [])
    return vd[0] if vd else None

def to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def fred_maps(defs, start="2000-01-01"):
    maps = {}
    for sid, label, fmt, freq in defs:
        obs = fred_obs(sid, observation_start=start)
        maps[sid] = {o["date"]: to_num(o["value"]) for o in obs if to_num(o["value"]) is not None}
        print(f"  {sid}: {len(maps[sid])} obs")
    return maps

def boc_observations(series_list, start="2000-01-01"):
    url = f"https://www.bankofcanada.ca/valet/observations/{','.join(series_list)}/json"
    r = requests.get(url, params={"start_date": start}, timeout=90)
    r.raise_for_status()
    obs = r.json().get("observations", [])
    maps = {s: {} for s in series_list}
    for row in obs:
        d = row.get("d")
        for s in series_list:
            cell = row.get(s)
            if cell and cell.get("v") not in (None, ""):
                maps[s][d] = to_num(cell["v"])
    return maps

# ---------- styles ----------
HDR_FILL   = PatternFill("solid", start_color="1F3864")
US_FILL    = PatternFill("solid", start_color="002868")
CA_FILL    = PatternFill("solid", start_color="8B0000")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=9)
TITLE_FONT = Font(name="Arial", bold=True, size=12)
META_FONT  = Font(name="Arial", italic=True, size=9, color="666666")
DATA_FONT  = Font(name="Arial", size=9)
ALT_FILL   = PatternFill("solid", start_color="EEF2F8")
POS_FONT   = Font(name="Arial", size=9, color="006100")
NEG_FONT   = Font(name="Arial", size=9, color="9C0006")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hcell(ws, r, c, v, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = HDR_FONT; cell.fill = fill or HDR_FILL
    cell.alignment = CENTER; cell.border = BORDER
    return cell

def dcell(ws, r, c, v, fill=None, fmt=None, font=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font or DATA_FONT
    cell.alignment = CENTER; cell.border = BORDER
    if fill: cell.fill = fill
    if fmt and isinstance(v, (int, float)): cell.number_format = fmt
    return cell

# ---------- vintage sheet ----------
def build_vintage_sheet(wb, series_id, title, sheet_name, numfmt, pulled_at):
    earliest = fred_earliest_vintage(series_id)
    start = earliest or "1947-01-01"
    print(f"  {sheet_name}: earliest vintage {earliest}")

    first = fred_obs(series_id, output_type=4, realtime_start="1776-07-04",
                     realtime_end="9999-12-31", observation_start=start)
    first_val = {o["date"]: to_num(o["value"]) for o in first}
    first_pub = {o["date"]: o["realtime_start"] for o in first}

    latest = fred_obs(series_id, observation_start=start)
    latest_val = {o["date"]: to_num(o["value"]) for o in latest}

    ws = wb.create_sheet(title=sheet_name)
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{title}  —  {series_id}"
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    ws["A2"] = (f"First-release vs latest-revised (ALFRED). Vintages begin {earliest}; "
                f"rows before that show the earliest archived value, not a true first print.  Pulled {pulled_at}")
    ws["A2"].font = META_FONT; ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 4

    for c, h in enumerate(["Obs Date", "First Release", "Latest Revised",
                           "Revision", "% Revision", "First Published"], 1):
        hcell(ws, 4, c, h)
    ws.row_dimensions[4].height = 26

    dates = sorted(set(first_val) | set(latest_val))
    row = 5
    for i, d in enumerate(dates):
        fv, lv = first_val.get(d), latest_val.get(d)
        fill = ALT_FILL if i % 2 == 0 else None
        # Derived values computed here (no LibreOffice available to recalc formulas)
        rev = (lv - fv) if (fv is not None and lv is not None) else None
        pct = (rev / abs(fv)) if (rev is not None and fv not in (None, 0)) else None
        rev_font = (POS_FONT if rev > 0 else NEG_FONT) if isinstance(rev, (int, float)) and rev != 0 else None
        dcell(ws, row, 1, d, fill=fill)
        dcell(ws, row, 2, fv, fill=fill, fmt=numfmt)
        dcell(ws, row, 3, lv, fill=fill, fmt=numfmt)
        dcell(ws, row, 4, rev if rev is not None else "", fill=fill, fmt=numfmt, font=rev_font)
        dcell(ws, row, 5, pct if pct is not None else "", fill=fill, fmt="0.0%", font=rev_font)
        dcell(ws, row, 6, first_pub.get(d, ""), fill=fill)
        ws.row_dimensions[row].height = 13
        row += 1

    for col, w in zip("ABCDEF", [12, 14, 14, 12, 11, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    print(f"    {len(dates)} observations")

# ---------- generic forward-filled panel ----------
def build_panel(wb, sheet_name, title, source_note, col_defs, data_maps, head_fill, pulled_at):
    """col_defs: list of (key, label, numfmt); data_maps: {key: {date: value}}"""
    ws = wb.create_sheet(title=sheet_name)
    ncol = 1 + len(col_defs)
    last_col = get_column_letter(ncol)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"{source_note}  Forward-filled across non-update days.  Pulled {pulled_at}"
    ws["A2"].font = META_FONT; ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 4

    hcell(ws, 4, 1, "Date")
    for c, (_, label, _) in enumerate(col_defs, start=2):
        hcell(ws, 4, c, label, fill=head_fill)
    ws.row_dimensions[4].height = 30

    all_dates = sorted({d for key, _, _ in col_defs for d in data_maps.get(key, {})})
    last = {key: None for key, _, _ in col_defs}
    row = 5
    for i, d in enumerate(all_dates):
        fill = ALT_FILL if i % 2 == 0 else None
        dcell(ws, row, 1, d, fill=fill)
        for c, (key, _, fmt) in enumerate(col_defs, start=2):
            v = data_maps.get(key, {}).get(d)
            if v is not None:
                last[key] = v
            dcell(ws, row, c, last[key] if last[key] is not None else "", fill=fill, fmt=fmt)
        ws.row_dimensions[row].height = 13
        row += 1

    ws.column_dimensions["A"].width = 12
    for c in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = "B5"
    print(f"  {sheet_name}: {len(all_dates)} rows x {len(col_defs)} series")

# ---------- main ----------
def main():
    pulled_at = datetime.now().strftime("%b %d, %Y %H:%M")

    # Load existing workbook (or start fresh) and drop any prior versions of our sheets
    if os.path.exists(WORKBOOK):
        print(f"Loading existing workbook: {WORKBOOK}")
        wb = load_workbook(WORKBOOK)
        for name in NEW_SHEETS:
            if name in wb.sheetnames:
                wb.remove(wb[name])
    else:
        print("No existing workbook found; creating new one.")
        wb = Workbook(); wb.remove(wb.active)

    print("\n=== FRED vintages (ALFRED) ===")
    for series_id, title, sheet_name, numfmt in VINTAGE:
        build_vintage_sheet(wb, series_id, title, sheet_name, numfmt, pulled_at)

    print("\n=== US financial conditions (FRED) ===")
    fc_maps = {}
    for sid, _, _ in FINCOND:
        obs = fred_obs(sid, observation_start="2000-01-01")
        fc_maps[sid] = {o["date"]: to_num(o["value"]) for o in obs if to_num(o["value"]) is not None}
        print(f"  {sid}: {len(fc_maps[sid])} obs")
    build_panel(wb, "US Financial Conditions",
                "US Financial Conditions & Rates — source: FRED (Federal Reserve)",
                "Daily/weekly/monthly series.", FINCOND, fc_maps, US_FILL, pulled_at)

    print("\n=== US SLOOS (FRED) ===")
    sloos_maps = fred_maps(FRED_SLOOS)
    build_panel(wb, "US SLOOS",
                "Senior Loan Officer Opinion Survey — source: FRED",
                "Quarterly. Net % of banks tightening C&I standards; positive = credit headwind.",
                [(s, l, f) for s, l, f, _ in FRED_SLOOS], sloos_maps, US_FILL, pulled_at)

    print("\n=== Commodity / input-cost prices (FRED) ===")
    com_maps = fred_maps(FRED_COMMODITY)
    build_panel(wb, "Commodity Prices",
                "Energy / Input-Cost Prices — source: FRED (US & Canada benchmarks)",
                "WTI & Henry Hub spot. Canadian WCS/AECO benchmarks require Alberta sources (not free via API).",
                [(s, l, f) for s, l, f, _ in FRED_COMMODITY], com_maps, US_FILL, pulled_at)

    print("\n=== Canada (Bank of Canada Valet) ===")
    boc_codes = [c for c, _, _ in BOC]
    boc_maps = boc_observations(boc_codes, start="2000-01-01")
    boc_defs = [(c, label, fmt) for c, label, fmt in BOC]
    for c in boc_codes:
        print(f"  {c}: {len(boc_maps.get(c, {}))} obs")
    # Computed GoC 3m-10Y term spread (mirrors the US FRED T10Y3M)
    ten = boc_maps.get("BD.CDN.10YR.DQ.YLD", {}); tb = boc_maps.get("TB.CDN.90D.MID", {})
    spread = {d: round(ten[d] - tb[d], 4) for d in ten if d in tb}
    boc_maps["GOC_3M10Y"] = spread
    boc_defs.append(("GOC_3M10Y", "GoC 3m-10Y Spread", "0.00"))
    print(f"  GoC 3m-10Y spread (computed): {len(spread)} obs")
    build_panel(wb, "Canada (BoC)",
                "Canada — source: Bank of Canada (Valet API)",
                "FX, policy/prime rates, GoC yields, 3M T-bill & computed 3m-10Y spread.", boc_defs, boc_maps, CA_FILL, pulled_at)

    # Reorder: Summary first, then the new free-data sheets, then everything else
    desired_front = ["Summary"] + NEW_SHEETS
    present_front = [s for s in desired_front if s in wb.sheetnames]
    rest = [s for s in wb.sheetnames if s not in present_front]
    order = present_front + rest
    wb._sheets.sort(key=lambda s: order.index(s.title))

    wb.save(WORKBOOK)
    print(f"\nSaved → {WORKBOOK}")

if __name__ == "__main__":
    main()
