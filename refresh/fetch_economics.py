"""
Trading Economics fetch + workbook build (CORRECTED).

Fixes two run-#2 regressions:
  1. Truncation — TE caps rows per API call. We now PACK batches by estimated row
     count (and URL length) so no call hits the cap, then VERIFY every series'
     earliest date against its true FirstValueDate and SOLO-RETRY any that come
     back short. Daily series capped at TE's hard 10,000-row limit are accepted.
  2. Dropped indicators — the original key indicators (Stock Market, Building
     Permits, tax rates) live outside the 8 category groups and were lost on
     rebuild. They are now always included.

Scope: 8 categories + Markets + Housing, plus the original key indicators.
Reuses cached country metadata in raw_te/ (no extra "latest" exports) and saves
every raw response to raw_te/ as an immutable layer.

Run fetch_fred_boc.py afterwards to re-add the free FRED/BoC sheets.
"""
import os, json, requests
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KEY = "cd31287336d24c4:7z6pdrvqoxp3fud"
RAW = "/Users/william.smith/Downloads/raw_te"
OUT = "/Users/william.smith/Downloads/economics_data.xlsx"
NOW = datetime.now()

CATEGORIES = ["GDP", "Labour", "Prices", "Money", "Trade", "Government",
              "Business", "Consumer", "Markets", "Housing"]
EXTRA_INDICATORS = ["Corporate Tax Rate", "Personal Income Tax Rate"]  # Taxes group, original keys
COUNTRY_API = {"US": "united%20states", "CA": "canada"}

KEY_INDICATORS = [
    "Stock Market", "GDP Growth Rate", "GDP Annual Growth Rate", "GDP Growth Annualized",
    "Unemployment Rate", "Inflation Rate", "Inflation Rate MoM", "Interest Rate",
    "Balance of Trade", "Current Account", "Current Account to GDP", "Government Debt to GDP",
    "Government Budget", "Business Confidence", "Manufacturing PMI", "Consumer Confidence",
    "Retail Sales MoM", "Building Permits", "Corporate Tax Rate", "Personal Income Tax Rate",
]
UNAVAILABLE = {
    "GDP Growth Annualized": "US: not a separately reported series (US GDP Growth Rate is already annualized)",
    "Manufacturing PMI":     "Requires a separate PMI-specific API endpoint",
}

CAP_ROWS, CAP_CHARS = 8000, 240

# ----------------------------------------------------------------- metadata / scope
def load_country(ctry):
    with open(f"{RAW}/country_{ctry}.json") as f:
        return json.load(f)

def build_scope(rows):
    """Return (latest_rows_in_scope, {indicator: meta}, ordered indicator list)."""
    latest, meta, names = [], {}, []
    want_extra = set(EXTRA_INDICATORS)
    for r in rows:
        cat = r.get("Category")
        grp = r.get("CategoryGroup")
        if grp in CATEGORIES or cat in want_extra:
            latest.append(r)
            meta[cat] = {"first": r.get("FirstValueDate"), "freq": (r.get("Frequency") or "").lower(),
                         "group": grp, "unit": r.get("Unit")}
            if cat not in names:
                names.append(cat)
    return latest, meta, names

# ----------------------------------------------------------------- cap-safe fetch
def est_rows(m):
    fvd = m.get("first")
    if not fvd:
        return 300
    try:
        yrs = max(0.5, (NOW - datetime.strptime(fvd[:10], "%Y-%m-%d")).days / 365.25)
    except Exception:
        return 300
    per = {"daily": 252, "weekly": 52, "monthly": 12, "quarterly": 4, "yearly": 1}.get(m.get("freq"), 12)
    est = yrs * per
    return int(min(est, 10000) if m.get("freq") == "daily" else est)

def pack(names, meta):
    batches, cur, rows, chars = [], [], 0, 0
    for n in names:
        e = est_rows(meta[n]); ln = len(n.replace(" ", "%20")) + 1
        if e >= CAP_ROWS or ln >= CAP_CHARS:
            if cur: batches.append(cur); cur, rows, chars = [], 0, 0
            batches.append([n]); continue
        if cur and (rows + e > CAP_ROWS or chars + ln > CAP_CHARS):
            batches.append(cur); cur, rows, chars = [], 0, 0
        cur.append(n); rows += e; chars += ln
    if cur: batches.append(cur)
    return batches

def fetch_batch(code, names):
    joined = ",".join(n.replace(" ", "%20") for n in names)
    url = f"https://api.tradingeconomics.com/historical/country/{code}/indicator/{joined}"
    r = requests.get(url, params={"c": KEY}, timeout=120)
    r.raise_for_status()
    return r.json()

def group_hist(rows):
    g = defaultdict(list)
    for r in rows:
        cat, dt, val = r.get("Category"), r.get("DateTime", "")[:10], r.get("Value")
        if cat and dt and val is not None:
            g[cat].append((dt, val))
    for k in g:
        g[k].sort(key=lambda x: x[0])
    return g

def truncated(ind, meta, dates):
    if not dates or not meta[ind].get("first"):
        return False
    try:
        earliest = min(dates)
        gap = (datetime.strptime(earliest, "%Y-%m-%d") - datetime.strptime(meta[ind]["first"][:10], "%Y-%m-%d")).days / 365.25
    except Exception:
        return False
    if gap <= 1.2:
        return False
    if meta[ind].get("freq") == "daily" and len(dates) >= 9900:  # TE 10k hard cap
        return False
    return True

def fetch_country_hist(ctry, names, meta):
    code = COUNTRY_API[ctry]
    batches = pack(names, meta)
    print(f"  {ctry}: {len(names)} indicators -> {len(batches)} cap-safe batches")
    all_rows = []
    for i, b in enumerate(batches, 1):
        rows = fetch_batch(code, b)
        all_rows.extend(rows)
        with open(f"{RAW}/{ctry}_batch_{i:03d}.json", "w") as f:
            json.dump(rows, f)
        print(f"    batch {i}/{len(batches)} ({len(b)} ind, {len(rows)} rows)", flush=True)
    hist = group_hist(all_rows)

    # verify + solo-retry
    retries = [n for n in names if truncated(n, meta, [d for d, _ in hist.get(n, [])])]
    if retries:
        print(f"  {ctry}: {len(retries)} came back short -> solo re-fetch: {retries[:8]}{'...' if len(retries)>8 else ''}")
    for n in retries:
        solo = fetch_batch(code, [n])
        with open(f"{RAW}/{ctry}_solo_{n.replace('/','-').replace(' ','_')[:40]}.json", "w") as f:
            json.dump(solo, f)
        hist[n] = group_hist(solo).get(n, [])
        if truncated(n, meta, [d for d, _ in hist[n]]):
            print(f"    WARN: {n} still short after solo fetch (hard cap or sparse series)")
    return hist

# ----------------------------------------------------------------- styling
HDR_FILL = PatternFill("solid", start_color="1F3864")
US_FILL  = PatternFill("solid", start_color="002868")
CA_FILL  = PatternFill("solid", start_color="8B0000")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
TITLE_FONT = Font(name="Arial", bold=True, size=12)
META_FONT = Font(name="Arial", italic=True, size=9, color="666666")
DATA_FONT = Font(name="Arial", size=9)
CAT_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
ALT_FILL = PatternFill("solid", start_color="EEF2F8")
CAT_FILLS = {
    "GDP": PatternFill("solid", start_color="1F3864"), "Labour": PatternFill("solid", start_color="1E5631"),
    "Prices": PatternFill("solid", start_color="7B1F1F"), "Money": PatternFill("solid", start_color="4A235A"),
    "Trade": PatternFill("solid", start_color="1A4A5A"), "Government": PatternFill("solid", start_color="5D4E37"),
    "Business": PatternFill("solid", start_color="1F3D1F"), "Consumer": PatternFill("solid", start_color="4A2040"),
    "Markets": PatternFill("solid", start_color="2C3E50"), "Housing": PatternFill("solid", start_color="6E2C00"),
}
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hcell(ws, r, c, v, fill=None, font=None, align=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font or HDR_FONT; cell.fill = fill or HDR_FILL
    cell.alignment = align or CENTER; cell.border = BORDER
    return cell

def dcell(ws, r, c, v, fill=None, align=CENTER):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = DATA_FONT; cell.alignment = align; cell.border = BORDER
    if fill: cell.fill = fill
    return cell

def fmt_date(d):
    try: return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%b %d, %Y")
    except: return d[:10] if d else ""

# ----------------------------------------------------------------- sheet builders
def build_summary(wb, us_latest, ca_latest, pulled_at):
    ws = wb.active; ws.title = "Summary"
    ws.merge_cells("A1:J1"); ws["A1"] = "Trading Economics — Full Indicator Summary (US & Canada)"
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = LEFT; ws.row_dimensions[1].height = 22
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Last pulled: {pulled_at}   |   US {len(us_latest)} / CA {len(ca_latest)} indicators   |   full history, cap-safe + verified"
    ws["A2"].font = META_FONT; ws["A2"].alignment = LEFT; ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 5
    us_map = {r["Category"]: r for r in us_latest}; ca_map = {r["Category"]: r for r in ca_latest}
    row = 4
    for cat in CATEGORIES + ["Taxes"]:
        us_cat = [r for r in us_latest if r.get("CategoryGroup") == cat]
        ca_cat = [r for r in ca_latest if r.get("CategoryGroup") == cat]
        all_inds = sorted(set(r["Category"] for r in us_cat) | set(r["Category"] for r in ca_cat))
        if not all_inds:
            continue
        ws.merge_cells(f"A{row}:J{row}")
        c = ws.cell(row=row, column=1, value=cat)
        c.font = CAT_FONT; c.fill = CAT_FILLS.get(cat, HDR_FILL); c.alignment = LEFT
        ws.row_dimensions[row].height = 16; row += 1
        for col, h in enumerate(["Indicator", "US Latest", "US Previous", "US Unit", "US Date",
                                  "CA Latest", "CA Previous", "CA Unit", "CA Date", "Frequency"], 1):
            hcell(ws, row, col, h)
        ws.row_dimensions[row].height = 16; row += 1
        for i, ind in enumerate(all_inds):
            fill = ALT_FILL if i % 2 == 0 else None
            us = us_map.get(ind, {}); ca = ca_map.get(ind, {})
            vals = [ind, us.get("LatestValue", ""), us.get("PreviousValue", ""), us.get("Unit", ""),
                    fmt_date(us.get("LatestValueDate", "")), ca.get("LatestValue", ""), ca.get("PreviousValue", ""),
                    ca.get("Unit", ""), fmt_date(ca.get("LatestValueDate", "")),
                    us.get("Frequency") or ca.get("Frequency") or ""]
            for col, v in enumerate(vals, 1):
                dcell(ws, row, col, v, fill=fill, align=LEFT if col == 1 else CENTER)
            ws.row_dimensions[row].height = 14; row += 1
        row += 1
    for col, w in enumerate([30, 12, 12, 14, 14, 12, 12, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"

def build_category_sheet(wb, cat, us_hist, ca_hist, us_latest, ca_latest):
    us_inds = sorted(set(r["Category"] for r in us_latest if r.get("CategoryGroup") == cat))
    ca_inds = sorted(set(r["Category"] for r in ca_latest if r.get("CategoryGroup") == cat))
    if not (us_inds or ca_inds):
        return
    ws = wb.create_sheet(title=cat)
    ws.merge_cells("A1:B1"); ws["A1"] = f"{cat} — Historical Data (Forward-filled)"
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = LEFT; ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 5
    both = sorted(set(us_inds) & set(ca_inds)); us_only = sorted(set(us_inds) - set(ca_inds)); ca_only = sorted(set(ca_inds) - set(us_inds))
    ordered = both + us_only + ca_only
    col_map = {}; cur = 2
    for ind in ordered:
        has_us, has_ca = ind in us_inds, ind in ca_inds
        uc = cur if has_us else None; cc = cur + (1 if has_us else 0) if has_ca else None
        col_map[ind] = (uc, cc); cur += (1 if has_us else 0) + (1 if has_ca else 0)
    hcell(ws, 3, 1, "Date"); ws.row_dimensions[3].height = 14; ws.row_dimensions[4].height = 48
    for ind in ordered:
        uc, cc = col_map[ind]
        if uc:
            hcell(ws, 3, uc, "United States", fill=US_FILL); hcell(ws, 4, uc, ind, fill=US_FILL)
            ws.column_dimensions[get_column_letter(uc)].width = 12
        if cc:
            hcell(ws, 3, cc, "Canada", fill=CA_FILL); hcell(ws, 4, cc, ind, fill=CA_FILL)
            ws.column_dimensions[get_column_letter(cc)].width = 12
    ws.column_dimensions["A"].width = 13; ws.freeze_panes = "A5"
    all_dates = sorted(set(d for ind in ordered for d, _ in us_hist.get(ind, [])) |
                       set(d for ind in ordered for d, _ in ca_hist.get(ind, [])))
    umap = {ind: dict(us_hist.get(ind, [])) for ind in ordered}
    cmap = {ind: dict(ca_hist.get(ind, [])) for ind in ordered}
    last = {ind: {"u": None, "c": None} for ind in ordered}
    for i, d in enumerate(all_dates):
        r = 5 + i; fill = ALT_FILL if i % 2 == 0 else None
        dcell(ws, r, 1, d, fill=fill)
        for ind in ordered:
            uc, cc = col_map[ind]
            if uc:
                v = umap[ind].get(d)
                if v is not None: last[ind]["u"] = v
                dcell(ws, r, uc, last[ind]["u"] if last[ind]["u"] is not None else "", fill=fill)
            if cc:
                v = cmap[ind].get(d)
                if v is not None: last[ind]["c"] = v
                dcell(ws, r, cc, last[ind]["c"] if last[ind]["c"] is not None else "", fill=fill)
        ws.row_dimensions[r].height = 13
    print(f"    {cat}: {len(all_dates)} dates x {len(ordered)} indicators")

def build_indicator_sheet(wb, ind, us_series, ca_series):
    ws = wb.create_sheet(title=ind[:31])
    ws.merge_cells("A1:C1"); ws["A1"] = ind
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = LEFT; ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 5
    note = UNAVAILABLE.get(ind)
    if note:
        ws.merge_cells("A3:C3"); ws["A3"] = f"Note: {note}"
        ws["A3"].font = META_FONT; ws["A3"].alignment = LEFT
        for col in "ABC": ws.column_dimensions[col].width = 16
        return
    for col, h in enumerate(["Date", "United States", "Canada"], 1):
        hcell(ws, 3, col, h)
    for col in "ABC": ws.column_dimensions[col].width = 16
    ws.freeze_panes = "A4"
    all_dates = sorted(set(d for d, _ in us_series) | set(d for d, _ in ca_series))
    um = dict(us_series); cm = dict(ca_series); lu = lc = None
    for d in all_dates:
        lu = um.get(d, lu); lc = cm.get(d, lc); um[d] = lu; cm[d] = lc
    for i, d in enumerate(all_dates):
        r = 4 + i; fill = ALT_FILL if i % 2 == 0 else None
        dcell(ws, r, 1, d, fill=fill); dcell(ws, r, 2, um.get(d, ""), fill=fill); dcell(ws, r, 3, cm.get(d, ""), fill=fill)
        ws.row_dimensions[r].height = 13

# ----------------------------------------------------------------- main
def main():
    pulled_at = NOW.strftime("%B %d, %Y  %H:%M")
    us_rows = load_country("US"); ca_rows = load_country("CA")
    us_latest, us_meta, us_names = build_scope(us_rows)
    ca_latest, ca_meta, ca_names = build_scope(ca_rows)
    print(f"Scope: US {len(us_names)} indicators, CA {len(ca_names)} indicators")

    print("\n=== Cap-safe historical fetch (US) ===")
    us_hist = fetch_country_hist("US", us_names, us_meta)
    print("\n=== Cap-safe historical fetch (CA) ===")
    ca_hist = fetch_country_hist("CA", ca_names, ca_meta)

    print("\n=== Building workbook ===")
    wb = Workbook()
    build_summary(wb, us_latest, ca_latest, pulled_at)
    for cat in CATEGORIES:
        build_category_sheet(wb, cat, us_hist, ca_hist, us_latest, ca_latest)
    for ind in KEY_INDICATORS:
        build_indicator_sheet(wb, ind, us_hist.get(ind, []), ca_hist.get(ind, []))
    wb.save(OUT)
    print(f"Saved -> {OUT}")
    print("Now run: python3 fetch_fred_boc.py   (re-adds the free FRED/BoC sheets)")

if __name__ == "__main__":
    main()
