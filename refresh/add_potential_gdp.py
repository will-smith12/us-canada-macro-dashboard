#!/usr/bin/env python3
"""
Add a "Potential GDP & Output Gap" sheet to master_economics_data.xlsx.

  US  - CBO potential GDP        : FRED GDPPOT (real, chained $B) + NGDPPOT (nominal)
  CAN - Bank of Canada output gap: Valet group INDINF_PRODUCT
        INDINF_OUTGAPMPR_Q  Current MPR output gap (%)   <- the published headline
        INDINF_OUTGAPR_Q    Historical / real-time MPR output gap (%)
        INDINF_OUTGAPI_Q    Output gap, Integrated framework (%)
        INDINF_OUTGAPM_Q    Output gap, Extended multivariate filter (%)

Quarterly; dates = quarter start (matches FRED + BoC + your FRED-vintage sheets).
House style: Date down col A, two-row header (row3 region, row4 series), data row5+.
Idempotent (replaces the sheet); refuses to run if the workbook is open.

Usage:  python3 add_potential_gdp.py [--offline]   (--offline uses already-downloaded files)
"""
from __future__ import annotations
import argparse, csv, datetime as dt, io, json, subprocess, urllib.request
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

TARGET = "master_economics_data.xlsx"
SHEET = "Potential GDP & Output Gap"
FRED = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={}"
VALET = ("https://www.bankofcanada.ca/valet/observations/"
         "INDINF_OUTGAPMPR_Q,INDINF_OUTGAPR_Q,INDINF_OUTGAPI_Q,INDINF_OUTGAPM_Q/json")

# column order: (region, label, key)
COLS = [
    ("United States", "Potential GDP - Real, CBO ($B, chained)",      "GDPPOT"),
    ("United States", "Potential GDP - Nominal, CBO ($B)",            "NGDPPOT"),
    ("Canada", "Output gap - Current MPR (%)",                        "INDINF_OUTGAPMPR_Q"),
    ("Canada", "Output gap - Historical / real-time MPR (%)",         "INDINF_OUTGAPR_Q"),
    ("Canada", "Output gap - Integrated framework (%)",              "INDINF_OUTGAPI_Q"),
    ("Canada", "Output gap - Extended multivariate filter (%)",      "INDINF_OUTGAPM_Q"),
]
PCT_KEYS = {c[2] for c in COLS if c[0] == "Canada"}


def get(url, offline_path=None, offline=False):
    if offline and offline_path and Path(offline_path).exists():
        return Path(offline_path).read_bytes()
    try:  # curl uses the system CA store (macOS python often can't verify TLS)
        return subprocess.run(["curl", "-sL", "--fail", url],
                              capture_output=True, check=True, timeout=120).stdout
    except Exception:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=60) as r:
            return r.read()


def to_qdate(s):  # 'YYYY-MM-DD' -> date
    return dt.date.fromisoformat(s[:10])


def load(offline=False):
    data = {}  # key -> {date: float}
    for fred_id, local in (("GDPPOT", "gdppot.csv"), ("NGDPPOT", "ngdppot.csv")):
        txt = get(FRED.format(fred_id), local, offline).decode()
        d = {}
        for row in csv.DictReader(io.StringIO(txt)):
            v = list(row.values())
            day, val = v[0], v[1]
            if val not in (".", "", None):
                d[to_qdate(day)] = float(val)
        data[fred_id] = d
    boc = json.loads(get(VALET, "boc_outgap.json", offline))
    for o in boc["observations"]:
        day = to_qdate(o["d"])
        for code in PCT_KEYS:
            if code in o and o[code].get("v") not in (None, ""):
                data.setdefault(code, {})[day] = float(o[code]["v"])
    return data


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--offline", action="store_true")
    ap.add_argument("--target", default=TARGET)
    args = ap.parse_args()

    tgt = Path(args.target)
    if (tgt.parent / f"~${tgt.name}").exists():
        raise SystemExit(f"ABORT: {tgt} is open in Excel (lock file). Close it first.")

    data = load(args.offline)
    dates = sorted({d for s in data.values() for d in s})
    today = dt.date.today().isoformat()
    print("series rows:", {k: len(v) for k, v in data.items()})
    print(f"date span: {dates[0]} -> {dates[-1]}  ({len(dates)} quarters)")

    wb = openpyxl.load_workbook(tgt)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    title_font, hdr_font = Font(bold=True, size=12), Font(bold=True)
    sub_font = Font(italic=True, color="555555")
    fill = PatternFill("solid", fgColor="DDEBF7")

    ws["A1"] = "Potential GDP (US, CBO) & Output Gap (Canada, Bank of Canada)"
    ws["A1"].font = title_font
    ws["A2"] = (f"Quarterly; dates = quarter start. US: FRED GDPPOT/NGDPPOT "
                f"(CBO estimate; includes forward projection). Canada: BoC Valet group "
                f"INDINF_PRODUCT. Pulled {today}.")
    ws["A2"].font = sub_font
    ws["A4"] = "Date"; ws["A4"].font = hdr_font
    for j, (region, label, key) in enumerate(COLS, start=2):
        cr = ws.cell(3, j, region); cr.font = hdr_font; cr.fill = fill
        cr.alignment = Alignment(horizontal="center")
        cl = ws.cell(4, j, label); cl.font = hdr_font
        cl.alignment = Alignment(wrap_text=True, vertical="top")
    for i, d in enumerate(dates, start=5):
        dc = ws.cell(i, 1, d); dc.number_format = "yyyy-mm-dd"
        for j, (region, label, key) in enumerate(COLS, start=2):
            v = data.get(key, {}).get(d)
            if v is not None:
                c = ws.cell(i, j, v)
                c.number_format = "0.0" if key in PCT_KEYS else "#,##0.0"
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 12
    for j in range(2, len(COLS) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 16

    # place right after the existing "GDP" sheet
    s = [x for x in wb._sheets if x.title != SHEET]
    obj = next(x for x in wb._sheets if x.title == SHEET)
    idx = next((i for i, x in enumerate(s) if x.title == "GDP"), len(s) - 1)
    wb._sheets = s[: idx + 1] + [obj] + s[idx + 1:]

    wb.save(tgt)
    print(f"Saved {tgt}: +{SHEET!r} ({len(COLS)} series x {len(dates)} quarters).")


if __name__ == "__main__":
    main()
