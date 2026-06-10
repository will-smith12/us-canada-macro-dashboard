import openpyxl, json, os

SRC = os.path.expanduser("~/Desktop/US_Canada_Macro_Indicators.xlsx")
wb = openpyxl.load_workbook(SRC, data_only=True)
data = wb["Data"]
meta = wb["Metadata"]

rows = list(data.iter_rows(values_only=True))
ind_row = rows[0]      # indicator names, span 2 cols
# build indicator -> (ca_col, us_col)
indicators = []
col = 1
while col < len(ind_row):
    name = ind_row[col]
    if name:
        indicators.append({"name": name, "ca_col": col, "us_col": col + 1})
    col += 2

# metadata lookup
mlookup = {}
for r in meta.iter_rows(min_row=2, values_only=True):
    name, country = r[0], r[1]
    if not name:
        continue
    mlookup[(name, country)] = {
        "unit": r[3], "frequency": r[4], "source": r[5],
        "observations": r[6], "latestDate": r[7], "latestValue": r[8],
    }

# per-indicator chart styling
CHART_CFG = {
    "GDP Growth Rate QoQ (%)":        {"type": "line",    "desc": "Quarter-over-quarter real GDP growth."},
    "GDP Growth Annualized (%)":      {"type": "line",    "desc": "Annualized quarterly GDP growth (US history is subscriber-gated)."},
    "GDP per Capita (USD)":           {"type": "line",    "desc": "Annual GDP per capita in current USD."},
    "Inflation Rate (%)":             {"type": "line",    "desc": "Year-over-year CPI inflation, monthly."},
    "Policy Interest Rate (%)":       {"type": "stepped", "desc": "Central bank policy rate."},
    "Labour Productivity (index)":    {"type": "line",    "desc": "Labour productivity index (points)."},
    "Gross Fixed Capital Formation":  {"type": "line",    "desc": "Investment in fixed assets. Note: Canada in CAD Million, US in USD Billion \u2014 plotted on separate axes.", "dualAxis": True},
    "GDP (USD B)":                    {"type": "line",    "desc": "Annual nominal GDP in current USD billions."},
    "Households Debt to GDP (%)":     {"type": "line",    "desc": "Household debt as a share of GDP."},
    "Households Debt to Income (%)":  {"type": "line",    "desc": "Household debt as a share of disposable income (US subscriber-gated)."},
    "Unemployment Rate (%)":          {"type": "line",    "desc": "Unemployment rate, monthly."},
}

out = {"generated": "2026-06-10", "source": "Trading Economics (compiled spreadsheet)", "indicators": []}

for ind in indicators:
    name = ind["name"]
    dates, ca_vals, us_vals = [], [], []
    for r in rows[2:]:
        d = r[0]
        if d is None:
            continue
        ca = r[ind["ca_col"]]
        us = r[ind["us_col"]]
        if ca is None and us is None:
            continue
        dates.append(str(d))
        ca_vals.append(ca)
        us_vals.append(us)
    cfg = CHART_CFG.get(name, {"type": "line", "desc": ""})
    ca_meta = mlookup.get((name, "Canada"), {})
    us_meta = mlookup.get((name, "United States"), {})

    # Skip indicators that don't have complete data: drop any tab where an
    # entire country series is missing (e.g. US subscriber-gated histories).
    ca_has = any(v is not None for v in ca_vals)
    us_has = any(v is not None for v in us_vals)
    if not (ca_has and us_has):
        print(f"  SKIP (incomplete): {name}")
        continue

    out["indicators"].append({
        "name": name,
        "chartType": cfg["type"],
        "dualAxis": cfg.get("dualAxis", False),
        "description": cfg["desc"],
        "unit": ca_meta.get("unit") or us_meta.get("unit") or "",
        "frequency": ca_meta.get("frequency") or us_meta.get("frequency") or "",
        "dates": dates,
        "canada": {"values": ca_vals, "unit": ca_meta.get("unit"), "source": ca_meta.get("source"),
                   "latestValue": ca_meta.get("latestValue"), "latestDate": str(ca_meta.get("latestDate") or "")},
        "us": {"values": us_vals, "unit": us_meta.get("unit"), "source": us_meta.get("source"),
               "latestValue": us_meta.get("latestValue"), "latestDate": str(us_meta.get("latestDate") or "")},
    })

with open("data.json", "w") as f:
    json.dump(out, f, indent=2, default=str)

# Also embed as JS so the dashboard works when opened directly (file://),
# where browsers block fetch() of local files.
with open("data.js", "w") as f:
    f.write("window.MACRO_DATA = " + json.dumps(out, default=str) + ";\n")

print("Wrote data.json + data.js with", len(out["indicators"]), "indicators")
for i in out["indicators"]:
    print(f"  {i['name']:35s} {i['chartType']:8s} pts={len(i['dates'])}")
