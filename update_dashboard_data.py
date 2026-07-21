#!/usr/bin/env python3
"""Rebuild the dashboard's data.json / data.js from the weekly-refreshed
``updating_master_macro_variables.xlsx`` workbook.

Design
------
* The existing ``data.json`` is used as a *template*: indicator names, chart
  type, dualAxis flag, descriptions, units and source labels are preserved
  verbatim (so the CONFLUENCE drawer links and the client-side GFCF
  CAD->USD conversion keep working). Only the numeric series + latest
  value/date are refreshed from the workbook.
* The workbook has two sheet shapes:
    - "long"  sheets: ``Date | United States | Canada`` (native cadence rows).
    - "wide"  category sheets: two-row header (country / sub-indicator),
      forward-filled to ~today on a weekly grid.
  Wide series are collapsed back to their true prints via change-point
  detection and the trailing forward-fill is truncated, so we never invent a
  not-yet-published current-period point.
* "Warranted" gate: the files are only rewritten when the extracted data
  actually changed (new dates or revised values); otherwise they are left
  untouched. Pass ``--force`` to always write, ``--dry-run`` to write the
  candidate to /tmp without touching the real files.

Must run with read access to ~/Downloads (i.e. under the weekly launchd job,
or a terminal with Full Disk Access).
"""
import argparse
import copy
import datetime
import json
import os
import sys

import openpyxl

WB_PATH = os.path.expanduser("~/Downloads/updating_master_macro_variables.xlsx")
DASH_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_JSON = os.path.join(DASH_DIR, "data.json")
DATA_JS = os.path.join(DASH_DIR, "data.js")
SOURCE_LABEL = "Trading Economics, FRED, Bank of Canada, StatCan (weekly refresh)"
START_YEAR = 2000
TODAY = datetime.date.today()
QUARTER_MONTHS = (3, 6, 9, 12)

# indicator name (exact data.json key) -> extraction spec
#   kind="long":  3-col sheet, US=col1 CA=col2, data from row 4
#   kind="wide":  category sheet, locate (country, sub) cols, data from row 5
#   cadence: monthly | quarterly | yearly | daily_m (daily resampled to monthly)
MAPPING = {
    "GDP Growth Rate QoQ (%)":       {"sheet": "GDP Growth Rate",  "kind": "long", "cadence": "quarterly"},
    "GDP per Capita (USD)":          {"sheet": "GDP",    "kind": "wide", "sub": "GDP per Capita", "cadence": "yearly"},
    "Inflation Rate (%)":            {"sheet": "Inflation Rate",   "kind": "long", "cadence": "monthly"},
    "Policy Interest Rate (%)":      {"sheet": "Interest Rate",    "kind": "long", "cadence": "daily_m"},
    "Labour Productivity (index)":   {"sheet": "Labour", "kind": "wide", "sub": "Productivity", "cadence": "quarterly"},
    "Gross Fixed Capital Formation": {"sheet": "GDP",    "kind": "wide", "sub": "Gross Fixed Capital Formation", "cadence": "quarterly"},
    "GDP (USD B)":                   {"sheet": "GDP",    "kind": "wide", "sub": "GDP", "cadence": "yearly"},
    "Households Debt to GDP (%)":    {"sheet": "Consumer", "kind": "wide", "sub": "Households Debt to GDP", "cadence": "quarterly"},
    "Unemployment Rate (%)":         {"sheet": "Unemployment Rate", "kind": "long", "cadence": "monthly"},
    # Small-business sentiment: US and CA live in the wide "Business" sheet under
    # different sub-indicator names (NFIB vs the CFIB "Small Business Sentiment"),
    # so this uses per-country sub keys instead of a single shared "sub".
    "Small Business Sentiment":      {"sheet": "Business", "kind": "wide",
                                      "sub_us": "NFIB Business Optimism Index",
                                      "sub_ca": "Small Business Sentiment",
                                      "cadence": "monthly"},
}


def _ym(dt):
    return f"{dt.year:04d}-{dt.month:02d}"


def _yy(dt):
    return f"{dt.year:04d}-01"


def _quarter_key(dt):
    qm = ((dt.month - 1) // 3) * 3 + 3
    return f"{dt.year:04d}-{qm:02d}"


def _parse_date(value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        return datetime.date(value.year, value.month, value.day)
    return datetime.date.fromisoformat(str(value)[:10])


def _raw_series(ws, col, data_start):
    """Return [(date, value)] for a column, dropping future-dated fill rows."""
    out = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        d = row[0]
        if d is None:
            continue
        try:
            dt = _parse_date(d)
        except (ValueError, TypeError):
            continue
        if dt > TODAY:
            continue
        v = row[col] if (col is not None and col < len(row)) else None
        out.append((dt, v))
    return out


def _wide_cols(ws, sub_us, sub_ca):
    """Locate the US and CA data columns in a wide category sheet.

    ``sub_us``/``sub_ca`` are the sub-indicator header strings for each country;
    pass the same value for both when a series shares one sub name across
    countries (the common case).
    """
    header = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    countries, subs = header[2], header[3]
    us = ca = None
    for ci in range(1, len(subs)):
        if subs[ci] is None:
            continue
        name = str(subs[ci]).strip()
        country = str(countries[ci]).strip() if ci < len(countries) else ""
        if name == sub_us and country.startswith("United"):
            us = ci
        elif name == sub_ca and country == "Canada":
            ca = ci
    return us, ca


def _resample_native(series, cadence):
    """Long sheets carry native-cadence rows; last value per period wins."""
    grid = {}
    for dt, v in series:
        if v is None or dt.year < START_YEAR:
            continue
        if cadence == "yearly":
            grid[_yy(dt)] = v
        elif cadence == "quarterly":
            if dt.month in QUARTER_MONTHS:
                grid[_ym(dt)] = v
        else:  # monthly / daily_m -> monthly bucket, last obs wins
            grid[_ym(dt)] = v
    return grid


def _resample_wide(series, cadence):
    """Forward-filled category column -> true prints on a cadence grid.

    Detect change points (first date of each new value), then step-fill onto
    the cadence grid up to the last real print so trailing weekly fill never
    creates a fake current-period point.
    """
    pts = [(dt, v) for dt, v in series if v is not None]
    if not pts:
        return {}
    change_points = []
    prev = object()
    for dt, v in pts:
        if v != prev:
            change_points.append((dt, v))
            prev = v

    def key(dt):
        if cadence == "yearly":
            return _yy(dt)
        if cadence == "quarterly":
            return _quarter_key(dt)
        return _ym(dt)

    cp_keyed = sorted((key(dt), v) for dt, v in change_points)
    last_key = key(change_points[-1][0])
    end_year = int(last_key[:4])

    periods = []
    for y in range(START_YEAR, end_year + 1):
        if cadence == "yearly":
            periods.append(f"{y:04d}-01")
        elif cadence == "quarterly":
            periods.extend(f"{y:04d}-{m:02d}" for m in QUARTER_MONTHS)
        else:
            periods.extend(f"{y:04d}-{m:02d}" for m in range(1, 13))
    periods = [p for p in periods if p <= last_key]

    grid = {}
    i, cur = 0, None
    for p in periods:
        while i < len(cp_keyed) and cp_keyed[i][0] <= p:
            cur = cp_keyed[i][1]
            i += 1
        if cur is not None:
            grid[p] = cur
    return grid


def _extract(wb, spec, log):
    sheet = spec["sheet"]
    if sheet not in wb.sheetnames:
        log.append(f"  WARN: sheet '{sheet}' not found; keeping existing data")
        return None
    ws = wb[sheet]
    cadence = spec["cadence"]
    if spec["kind"] == "long":
        us = _resample_native(_raw_series(ws, 1, 4), cadence)
        ca = _resample_native(_raw_series(ws, 2, 4), cadence)
    else:
        sub_us = spec.get("sub_us", spec.get("sub"))
        sub_ca = spec.get("sub_ca", spec.get("sub"))
        us_col, ca_col = _wide_cols(ws, sub_us, sub_ca)
        if us_col is None and ca_col is None:
            wanted = sub_us if sub_us == sub_ca else f"{sub_us!r}/{sub_ca!r}"
            log.append(f"  WARN: sub '{wanted}' not found in '{sheet}'; keeping existing data")
            return None
        us = _resample_wide(_raw_series(ws, us_col, 5), cadence)
        ca = _resample_wide(_raw_series(ws, ca_col, 5), cadence)
    return us, ca


def _last_non_null(values, dates):
    for v, d in zip(reversed(values), reversed(dates)):
        if v is not None:
            return v, d
    return None, ""


def build(template, wb, log):
    out = copy.deepcopy(template)
    out["generated"] = TODAY.isoformat()
    out["source"] = SOURCE_LABEL
    for ind in out["indicators"]:
        name = ind["name"]
        spec = MAPPING.get(name)
        if spec is None:
            log.append(f"  WARN: no mapping for '{name}'; keeping existing data")
            continue
        extracted = _extract(wb, spec, log)
        if extracted is None:
            continue
        us_grid, ca_grid = extracted

        # Preserve-history merge: the new source takes precedence wherever it
        # has a value, but any period it lacks is back-filled from the existing
        # series so we never drop points the dashboard already showed (e.g. the
        # CA policy rate before 2009, absent from the new workbook's sheet).
        old_ca = dict(zip(ind["dates"], ind["canada"]["values"]))
        old_us = dict(zip(ind["dates"], ind["us"]["values"]))
        for k, v in old_ca.items():
            if v is not None and ca_grid.get(k) is None:
                ca_grid[k] = v
        for k, v in old_us.items():
            if v is not None and us_grid.get(k) is None:
                us_grid[k] = v

        keys = sorted(set(us_grid) | set(ca_grid))
        dates, ca_vals, us_vals = [], [], []
        for k in keys:
            cv, uv = ca_grid.get(k), us_grid.get(k)
            if cv is None and uv is None:
                continue
            dates.append(k)
            ca_vals.append(cv)
            us_vals.append(uv)
        if not dates:
            log.append(f"  WARN: '{name}' produced no data; keeping existing data")
            continue
        ind["dates"] = dates
        ind["canada"]["values"] = ca_vals
        ind["us"]["values"] = us_vals
        cval, cdate = _last_non_null(ca_vals, dates)
        uval, udate = _last_non_null(us_vals, dates)
        ind["canada"]["latestValue"] = cval
        ind["canada"]["latestDate"] = cdate
        ind["us"]["latestValue"] = uval
        ind["us"]["latestDate"] = udate
        log.append(f"  {name:32s} n={len(dates):4d}  end={dates[-1]}  "
                   f"CA={cval}  US={uval}")
    return out


def _normalized(obj):
    o = dict(obj)
    o.pop("generated", None)
    return json.dumps(o, sort_keys=True, default=str)


def _changes(old, new):
    old_by = {i["name"]: i for i in old.get("indicators", [])}
    rows = []
    for ind in new["indicators"]:
        o = old_by.get(ind["name"])
        if o is None:
            rows.append(f"  + NEW  {ind['name']}")
            continue
        if (o.get("dates") != ind.get("dates")
                or o.get("canada", {}).get("values") != ind["canada"]["values"]
                or o.get("us", {}).get("values") != ind["us"]["values"]):
            old_end = o.get("dates", [""])[-1] if o.get("dates") else "-"
            new_end = ind["dates"][-1] if ind["dates"] else "-"
            rows.append(f"  ~ {ind['name']:32s} {old_end} -> {new_end}")
    return rows


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", default=WB_PATH)
    ap.add_argument("--force", action="store_true", help="write even if unchanged")
    ap.add_argument("--dry-run", action="store_true",
                    help="write candidate to /tmp only; never touch the real files")
    ap.add_argument("--report", default=None, help="path to also write the run report")
    args = ap.parse_args()

    if not os.path.exists(DATA_JSON):
        print(f"ERROR: template {DATA_JSON} missing", file=sys.stderr)
        return 2
    if not os.path.exists(args.workbook):
        print(f"ERROR: workbook {args.workbook} not readable", file=sys.stderr)
        return 2

    with open(DATA_JSON) as f:
        template = json.load(f)
    baseline = copy.deepcopy(template)

    wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    log = []
    out = build(template, wb, log)
    wb.close()

    changed = _normalized(out) != _normalized(baseline)
    change_rows = _changes(baseline, out)

    report = [f"update_dashboard_data {datetime.datetime.now():%Y-%m-%d %H:%M:%S}",
              f"workbook: {args.workbook}",
              f"changed: {changed}", ""]
    report += log + [""]
    report.append("Indicators changed vs current data.json:" if change_rows
                   else "No indicator series changed.")
    report += change_rows
    text = "\n".join(report)
    print(text)
    if args.report:
        with open(args.report, "w") as f:
            f.write(text + "\n")

    if args.dry_run:
        cand = "/tmp/dash_candidate.json"
        with open(cand, "w") as f:
            json.dump(out, f, indent=2, default=str)
        print(f"\n[dry-run] candidate written to {cand} (real files untouched)")
        return 0

    if not changed and not args.force:
        print("\nNo change -> data.json / data.js left untouched.")
        return 0

    with open(DATA_JSON, "w") as f:
        json.dump(out, f, indent=2, default=str)
    with open(DATA_JS, "w") as f:
        f.write("window.MACRO_DATA = " + json.dumps(out, default=str) + ";\n")
    print(f"\nWrote {DATA_JSON} + {DATA_JS} ({len(out['indicators'])} indicators).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
