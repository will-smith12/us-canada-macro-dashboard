#!/usr/bin/env python3
"""
Build a clean, human-readable WIDE workbook of BoC Business Outlook Survey results,
one formatted sheet per indicator, with sectors and regions shown side by side.

Reads the tidy sheets from data/BoC_BOS_sector_region.xlsx (no network needed).
Output: bos/data/BoC_BOS_wide_readable.xlsx
"""
from __future__ import annotations
import os
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.environ.get("BOS_XLSX") or os.path.join(HERE, "data", "BoC_BOS_sector_region.xlsx")
OUT = os.environ.get("BOS_WIDE_XLSX") or os.path.join(HERE, "data", "BoC_BOS_wide_readable.xlsx")

# ---- member display (short) + order ------------------------------------
SEC_SHORT = {
    "Primary": "Primary", "Manufacturing": "Manufacturing",
    "CITU (constr./info/transp./util.)": "CITU", "Trade": "Trade",
    "FIRE (finance/insur./real estate)": "FIRE",
    "CPBS (comm./pers./bus. services)": "CPBS",
}
SEC_ORDER = ["Primary", "Manufacturing", "CITU", "Trade", "FIRE", "CPBS"]
REG_SHORT = {
    "Atlantic": "Atlantic", "Quebec": "Quebec", "Ontario": "Ontario",
    "Prairies": "Prairies", "British Columbia": "BC",
    "All regions (indicator)": "All regions",
}
REG_ORDER = ["Atlantic", "Quebec", "Ontario", "Prairies", "BC"]

# ---- sheet spec: (sheet_name, indicator, subcomponent|None, unit) -------
SPECS = [
    ("Past sales growth",          "Past sales growth", None, "Balance of opinion"),
    ("Past sales declines",        "Past sales declines", "Share of firms reporting declines", "% of firms"),
    ("Future sales",               "Future sales growth", "Future sales (balance of opinion)", "Balance of opinion"),
    ("Future sales indicators",    "Future sales growth", "Indicators of future sales", "Balance of opinion"),
    ("Investment M&E",             "Investment in machinery & equipment", None, "Balance of opinion"),
    ("Credit conditions",          "Credit conditions", None, "Balance of opinion"),
    ("Employment",                 "Employment", None, "Balance of opinion"),
    ("Capacity some difficulty",   "Capacity pressures", "Some difficulty meeting demand", "% of firms"),
    ("Capacity signif difficulty", "Capacity pressures", "Significant difficulty meeting demand", "% of firms"),
    ("Labour shortages",           "Labour shortages", "Share reporting shortages", "% of firms"),
    ("Labour shortage intensity",  "Labour shortage intensity", None, "Balance of opinion"),
    ("Wages",                      "Wages", None, "Balance of opinion"),
    ("Input prices",               "Input prices", None, "Balance of opinion"),
    ("Output prices",              "Output prices", None, "Balance of opinion"),
    ("Inflation exp below 1pct",   "Inflation expectations (next 2 years)", "Below 1%", "% of firms"),
    ("Inflation exp 1 to 2pct",    "Inflation expectations (next 2 years)", "1% to 2%", "% of firms"),
    ("Inflation exp 2 to 3pct",    "Inflation expectations (next 2 years)", "2% to 3%", "% of firms"),
    ("Inflation exp above 3pct",   "Inflation expectations (next 2 years)", "Above 3%", "% of firms"),
    ("Inflation exp no response",  "Inflation expectations (next 2 years)", "No response", "% of firms"),
]
REGIND = ("Regional BOS indicator", "Regional BOS indicator",
          "Contribution to regional indicator", "Standardized units")

# ---- styling -----------------------------------------------------------
NAVY = "1F3864"
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUB_FONT = Font(italic=True, size=10, color="595959")
QHDR_FONT = Font(bold=True, size=10, color="FFFFFF")
SEC_GRP_FILL = PatternFill("solid", fgColor="2E5597")   # blue
REG_GRP_FILL = PatternFill("solid", fgColor="548235")   # green
GRP_FONT = Font(bold=True, size=11, color="FFFFFF")
SEC_HDR_FILL = PatternFill("solid", fgColor="D9E1F2")   # light blue
REG_HDR_FILL = PatternFill("solid", fgColor="E2EFDA")   # light green
Q_HDR_FILL = PatternFill("solid", fgColor="404040")     # dark gray
COLHDR_FONT = Font(bold=True, size=10, color="1F1F1F")
BAND_FILL = PatternFill("solid", fgColor="F5F8FC")      # subtle band
QCELL_FONT = Font(bold=True, size=10, color="404040")

thin = Side(style="thin", color="D0D0D0")
med = Side(style="medium", color="808080")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def qkey(q: str) -> int:
    return int(q[:4]) * 4 + int(q[5])


def qlabel(q: str) -> str:
    return f"{q[:4]} Q{q[5]}"


def pivot_block(df, indicator, sub, short_map, order):
    m = df[df["indicator"] == indicator]
    if sub is None:
        m = m[m["subcomponent"].isna() | (m["subcomponent"].astype(str).str.strip() == "")]
    else:
        m = m[m["subcomponent"] == sub]
    if m.empty:
        return pd.DataFrame()
    p = m.pivot_table(index="quarter", columns="member", values="value", aggfunc="first")
    p.columns = [short_map.get(c, c) for c in p.columns]
    cols = [c for c in order if c in p.columns]
    return p[cols]


def build_sheet(wb, sec_df, reg_df, spec, region_only=False):
    name, indicator, sub, unit = spec
    ws = wb.create_sheet(title=name[:31])

    sec = pd.DataFrame() if region_only else pivot_block(sec_df, indicator, sub, SEC_SHORT, SEC_ORDER)
    if name == "Regional BOS indicator":
        reg = pivot_block(reg_df, indicator, sub, REG_SHORT, REG_ORDER + ["All regions"])
    else:
        reg = pivot_block(reg_df, indicator, sub, REG_SHORT, REG_ORDER)

    quarters = sorted(set(list(sec.index) + list(reg.index)), key=qkey)
    sec_cols = list(sec.columns)
    reg_cols = list(reg.columns)

    # ---- header block ----
    title = indicator + (f" — {sub}" if sub else "")
    ncols = 1 + len(sec_cols) + len(reg_cols)
    last = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT
    cov = f"{qlabel(quarters[0])} – {qlabel(quarters[-1])}" if quarters else ""
    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = (f"Units: {unit}   •   Quarterly, four-quarter moving average   •   "
                f"Coverage: {cov}   •   Source: Bank of Canada Valet API (BOS)")
    ws["A2"].font = SUB_FONT
    ws["A2"].alignment = LEFT
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 15

    grp_row, hdr_row, first_data = 4, 5, 6

    # Quarter header (merged over the two header rows)
    ws.merge_cells(start_row=grp_row, start_column=1, end_row=hdr_row, end_column=1)
    qc = ws.cell(row=grp_row, column=1, value="Quarter")
    qc.font = QHDR_FONT
    qc.fill = Q_HDR_FILL
    qc.alignment = CENTER

    col = 2
    sec_start = col
    if sec_cols:
        ws.merge_cells(start_row=grp_row, start_column=col, end_row=grp_row,
                       end_column=col + len(sec_cols) - 1)
        g = ws.cell(row=grp_row, column=col, value="By sector")
        g.font = GRP_FONT
        g.fill = SEC_GRP_FILL
        g.alignment = CENTER
        for c in sec_cols:
            cell = ws.cell(row=hdr_row, column=col, value=c)
            cell.font = COLHDR_FONT
            cell.fill = SEC_HDR_FILL
            cell.alignment = CENTER_WRAP
            col += 1
    sec_end = col - 1
    reg_start = col
    if reg_cols:
        grp_label = "By region — contributions" if name == "Regional BOS indicator" else "By region"
        ws.merge_cells(start_row=grp_row, start_column=col, end_row=grp_row,
                       end_column=col + len(reg_cols) - 1)
        g = ws.cell(row=grp_row, column=col, value=grp_label)
        g.font = GRP_FONT
        g.fill = REG_GRP_FILL
        g.alignment = CENTER
        for c in reg_cols:
            cell = ws.cell(row=hdr_row, column=col, value=c)
            cell.font = COLHDR_FONT
            cell.fill = REG_HDR_FILL
            cell.alignment = CENTER_WRAP
            col += 1
    reg_end = col - 1

    # ---- data ----
    for i, q in enumerate(quarters):
        r = first_data + i
        band = (i % 2 == 1)
        qcell = ws.cell(row=r, column=1, value=qlabel(q))
        qcell.font = QCELL_FONT
        qcell.alignment = CENTER
        if band:
            qcell.fill = BAND_FILL
        cidx = 2
        for c in sec_cols:
            v = sec.loc[q, c] if (not sec.empty and q in sec.index) else None
            cell = ws.cell(row=r, column=cidx)
            cell.value = None if pd.isna(v) else float(v)
            cell.number_format = "0.0"
            cell.alignment = RIGHT
            if band:
                cell.fill = BAND_FILL
            cidx += 1
        for c in reg_cols:
            v = reg.loc[q, c] if (not reg.empty and q in reg.index) else None
            cell = ws.cell(row=r, column=cidx)
            cell.value = None if pd.isna(v) else float(v)
            cell.number_format = "0.0"
            cell.alignment = RIGHT
            if band:
                cell.fill = BAND_FILL
            cidx += 1

    # ---- borders (thin grid on header+data; medium block separators) ----
    last_row = first_data + len(quarters) - 1
    for r in range(grp_row, last_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # medium vertical separators after Quarter and between sector/region
    for r in range(grp_row, last_row + 1):
        a = ws.cell(row=r, column=1)
        a.border = Border(left=a.border.left, right=med, top=a.border.top, bottom=a.border.bottom)
        if reg_cols and sec_cols:
            b = ws.cell(row=r, column=sec_end)
            b.border = Border(left=b.border.left, right=med, top=b.border.top, bottom=b.border.bottom)

    # ---- widths + freeze ----
    ws.column_dimensions["A"].width = 9
    for c in range(2, ncols + 1):
        hdr = ws.cell(row=hdr_row, column=c).value or ""
        ws.column_dimensions[get_column_letter(c)].width = max(len(str(hdr)) + 2, 10)
    ws.freeze_panes = "B6"
    ws.sheet_view.showGridLines = False


def build_contents(wb, specs_done):
    ws = wb.create_sheet(title="Contents", index=0)
    ws["A1"] = "Bank of Canada Business Outlook Survey — readable wide tables"
    ws["A1"].font = Font(bold=True, size=15, color=NAVY)
    ws.merge_cells("A1:E1")
    ws["A2"] = ("One sheet per indicator. Quarters run down the side; sector and region "
                "breakdowns are shown side by side. Values are four-quarter moving averages.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:E2")

    heads = ["Sheet", "Indicator", "Breakdown / subcomponent", "Unit", "Coverage"]
    hr = 4
    for j, h in enumerate(heads, start=1):
        c = ws.cell(row=hr, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E5597")
        c.alignment = CENTER
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, (name, ind, sub, unit, cov) in enumerate(specs_done):
        r = hr + 1 + i
        vals = [name, ind, sub or "—", unit, cov]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = LEFT if j <= 3 else CENTER
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if i % 2:
                c.fill = BAND_FILL

    # legend
    lr = hr + 2 + len(specs_done)
    ws.cell(row=lr, column=1, value="Sector codes (NAICS)").font = Font(bold=True, color=NAVY)
    legend = [
        ("Primary", "11, 21"), ("Manufacturing", "311–339"),
        ("CITU", "Construction, information, transportation, utilities (22, 23, 48, 49, 51)"),
        ("Trade", "41, 44, 45"), ("FIRE", "Finance, insurance, real estate (52–53)"),
        ("CPBS", "Commercial, personal & business services (54, 55, 56, 71, 72, 81)"),
    ]
    for k, (code, desc) in enumerate(legend, start=1):
        ws.cell(row=lr + k, column=1, value=code).font = Font(bold=True)
        ws.cell(row=lr + k, column=2, value=desc)
    rr = lr + len(legend) + 2
    ws.cell(row=rr, column=1, value="Regions").font = Font(bold=True, color=NAVY)
    ws.cell(row=rr + 1, column=1, value="AT=Atlantic, QC=Quebec, ON=Ontario, "
            "PR=Prairies, BC=British Columbia")
    cr = rr + 3
    notes = [
        ("Source", "Bank of Canada, Business Outlook Survey (Valet API). Public, no key."),
        ("Retrieved", date.today().isoformat()),
        ("Units", "'Balance of opinion' = % reporting higher/more minus % lower/less (may be "
                  "negative). '% of firms' = share of respondents. 'Standardized units' = "
                  "regional indicator scale."),
        ("Caveat", "Four-quarter moving averages may not match the aggregate BOS figures "
                   "published each quarter."),
        ("Note", "Past sales growth question was dropped from the survey in 2023Q1 (blank after)."),
        ("Full dataset", "See BoC_BOS_sector_region.xlsx for tidy long tables, firm-size "
                         "breakdowns, and full-precision values."),
    ]
    for k, (a, b) in enumerate(notes):
        ws.cell(row=cr + k, column=1, value=a).font = Font(bold=True)
        cell = ws.cell(row=cr + k, column=2, value=b)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.sheet_view.showGridLines = False


def main():
    sec_df = pd.read_excel(SRC, sheet_name="Sector_tidy")
    reg_df = pd.read_excel(SRC, sheet_name="Region_tidy")

    wb = Workbook()
    wb.remove(wb.active)

    specs_done = []
    for spec in SPECS:
        build_sheet(wb, sec_df, reg_df, spec)
        name, ind, sub, unit = spec
        # coverage for contents
        block = pivot_block(sec_df, ind, sub, SEC_SHORT, SEC_ORDER)
        if block.empty:
            block = pivot_block(reg_df, ind, sub, REG_SHORT, REG_ORDER)
        qs = sorted(block.index, key=qkey)
        cov = f"{qlabel(qs[0])} – {qlabel(qs[-1])}" if len(qs) else ""
        specs_done.append((name, ind, sub, unit, cov))

    # regional indicator (region only)
    build_sheet(wb, sec_df, reg_df, REGIND, region_only=True)
    rb = pivot_block(reg_df, REGIND[1], REGIND[2], REG_SHORT, REG_ORDER + ["All regions"])
    qs = sorted(rb.index, key=qkey)
    specs_done.append((REGIND[0], REGIND[1], REGIND[2], REGIND[3],
                       f"{qlabel(qs[0])} – {qlabel(qs[-1])}" if len(qs) else ""))

    build_contents(wb, specs_done)
    # order: Contents first (already index 0), keep rest as added
    wb.save(OUT)
    print("Wrote", OUT)
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
