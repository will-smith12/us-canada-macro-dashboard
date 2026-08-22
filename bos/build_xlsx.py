#!/usr/bin/env python3
"""
Build a standalone spreadsheet of the Bank of Canada Business Outlook Survey (BOS)
*disaggregated* series (by sector, by region, by firm size) from the public BoC
Valet API.

BoC publishes standing panels of BOS results broken down by:
  - sector   -> groups BOS_BYSECTOR_*   (6 sectors: Primary, Manufacturing, CITU, Trade, FIRE, CPBS)
  - region   -> groups BOS_BYREGION_*   (5 regions: Atlantic, Quebec, Ontario, Prairies, BC)
  - firm size-> groups BOS_BYSIZE_*      (3 sizes: Small, Medium-sized, Large)
plus a regional summary indicator BOS_REGIONAL_INDICATOR (contributions by region).

All disaggregated series are quarterly FOUR-QUARTER MOVING AVERAGES.

Output: bos/data/BoC_BOS_sector_region.xlsx
"""
from __future__ import annotations
import json
import os
import subprocess
import sys
import time
from datetime import date, datetime

import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(HERE, "cache")
CANDIDATES = os.path.join(HERE, "candidates.json")
OUT_XLSX = os.environ.get("BOS_XLSX") or os.path.join(HERE, "data", "BoC_BOS_sector_region.xlsx")
VALET = "https://www.bankofcanada.ca/valet"
UA = "Mozilla/5.0 (macOS) bos-harvest/1.0"

os.makedirs(CACHE, exist_ok=True)

# ---------------------------------------------------------------------------
# Parsing maps
# ---------------------------------------------------------------------------
# indicator code prefix -> (indicator, subcomponent, unit)
PREFIX_MAP = {
    "PS":       ("Past sales growth", "", "Balance of opinion"),
    "PSD":      ("Past sales declines", "Share of firms reporting declines", "% of firms"),
    "FS":       ("Future sales growth", "Future sales (balance of opinion)", "Balance of opinion"),
    "FSI":      ("Future sales growth", "Indicators of future sales", "Balance of opinion"),
    "INV":      ("Investment in machinery & equipment", "", "Balance of opinion"),
    "EMP":      ("Employment", "", "Balance of opinion"),
    "FIN":      ("Credit conditions", "", "Balance of opinion"),
    "CAPSOME":  ("Capacity pressures", "Some difficulty meeting demand", "% of firms"),
    "CAPSIGN":  ("Capacity pressures", "Significant difficulty meeting demand", "% of firms"),
    "LS":       ("Labour shortages", "Share reporting shortages", "% of firms"),
    "LSI":      ("Labour shortage intensity", "", "Balance of opinion"),
    "WAGE":     ("Wages", "", "Balance of opinion"),
    "INP":      ("Input prices", "", "Balance of opinion"),
    "OUP":      ("Output prices", "", "Balance of opinion"),
    "INF_BELOW_1":    ("Inflation expectations (next 2 years)", "Below 1%", "% of firms"),
    "INF_BETWEEN1_2": ("Inflation expectations (next 2 years)", "1% to 2%", "% of firms"),
    "INF_BETWEEN2_3": ("Inflation expectations (next 2 years)", "2% to 3%", "% of firms"),
    "INF_ABOVE_3":    ("Inflation expectations (next 2 years)", "Above 3%", "% of firms"),
    "INFNA":          ("Inflation expectations (next 2 years)", "No response", "% of firms"),
    "REGIONAL": ("Regional BOS indicator", "Contribution to regional indicator", "Standardized units"),
}
# stable ordering for indicators in output
INDICATOR_ORDER = [
    "Past sales growth", "Past sales declines", "Future sales growth",
    "Investment in machinery & equipment", "Employment", "Capacity pressures",
    "Labour shortages", "Labour shortage intensity", "Wages",
    "Input prices", "Output prices", "Credit conditions",
    "Inflation expectations (next 2 years)", "Regional BOS indicator",
]

SECTOR_MEMBERS = {
    "PRIMARY": "Primary", "MANUFACT": "Manufacturing",
    "CITU": "CITU (constr./info/transp./util.)", "TRADE": "Trade",
    "FIRE": "FIRE (finance/insur./real estate)",
    "CPBS": "CPBS (comm./pers./bus. services)",
}
SECTOR_ORDER = ["PRIMARY", "MANUFACT", "CITU", "TRADE", "FIRE", "CPBS"]

REGION_MEMBERS = {"AT": "Atlantic", "QC": "Quebec", "ON": "Ontario",
                  "PR": "Prairies", "BC": "British Columbia"}
REGION_ORDER = ["AT", "QC", "ON", "PR", "BC"]

SIZE_MEMBERS = {"S": "Small", "M": "Medium-sized", "L": "Large"}
SIZE_ORDER = ["S", "M", "L"]

REGIND_MEMBERS = {"ATLANTIC": "Atlantic", "QC": "Quebec", "ON": "Ontario",
                  "PRAIRIES": "Prairies", "BC": "British Columbia",
                  "INDICATOR": "All regions (indicator)"}
REGIND_ORDER = ["ATLANTIC", "QC", "ON", "PRAIRIES", "BC", "INDICATOR"]


def dim_for_group(code: str):
    if code.startswith("BOS_BYSECTOR_"):
        return "Sector", SECTOR_MEMBERS, SECTOR_ORDER
    if code.startswith("BOS_BYREGION_"):
        return "Region", REGION_MEMBERS, REGION_ORDER
    if code.startswith("BOS_BYSIZE_"):
        return "Firm size", SIZE_MEMBERS, SIZE_ORDER
    if code == "BOS_REGIONAL_INDICATOR":
        return "Region", REGIND_MEMBERS, REGIND_ORDER
    return None, None, None


def split_code(series_code: str, members: dict):
    """Return (indicator_prefix, member_key) by stripping the longest matching
    dimension member suffix from the series code."""
    body = series_code[4:] if series_code.startswith("BOS_") else series_code
    best = None
    for msuf in members:
        if body == msuf or body.endswith("_" + msuf):
            if best is None or len(msuf) > len(best):
                best = msuf
    if best is None:
        return None, None
    prefix = "" if body == best else body[: -(len(best) + 1)]
    return prefix, best


# ---------------------------------------------------------------------------
# Fetch
# ---------------------------------------------------------------------------
def _get(url: str, tries: int = 4):
    last = None
    for i in range(tries):
        try:
            out = subprocess.run(
                ["curl", "-sS", "--fail", "--max-time", "60", "-A", UA, url],
                capture_output=True, text=True, check=True,
            )
            return json.loads(out.stdout)
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(1.2 * (i + 1))
    raise RuntimeError(f"fetch failed: {url}: {last}")


def fetch_group(code: str) -> dict:
    cache_path = os.path.join(CACHE, f"{code}.json")
    if os.path.exists(cache_path):
        with open(cache_path) as fh:
            return json.load(fh)
    url = f"{VALET}/observations/group/{code}/json"
    data = _get(url)
    with open(cache_path, "w") as fh:
        json.dump(data, fh)
    time.sleep(0.25)
    return data


def target_groups() -> list[str]:
    codes = [c["code"] for c in json.load(open(CANDIDATES))]
    targets = [
        c for c in codes
        if c.startswith(("BOS_BYSECTOR_", "BOS_BYREGION_", "BOS_BYSIZE_"))
    ]
    if "BOS_REGIONAL_INDICATOR" in codes:
        targets.append("BOS_REGIONAL_INDICATOR")
    return sorted(set(targets))


# ---------------------------------------------------------------------------
# Parse
# ---------------------------------------------------------------------------
def parse_all():
    rows = []
    catalog = []
    definitions = {}   # indicator prefix -> example series description
    unparsed = []

    for code in target_groups():
        data = fetch_group(code)
        dim, members, order = dim_for_group(code)
        gd = data.get("groupDetail", {})
        sdet = data.get("seriesDetail", {})
        obs = data.get("observations", [])
        page_link = gd.get("link", "")
        valet_url = f"{VALET}/observations/group/{code}/json"
        gdesc = gd.get("description", "")

        g_inds, g_members, g_units = set(), set(), set()
        g_dates = []
        for o in obs:
            d = o.get("d")
            for scode, sdd in sdet.items():
                cell = o.get(scode)
                if not cell:
                    continue
                v = cell.get("v")
                if v in (None, "", "NaN"):
                    continue
                prefix, mkey = split_code(scode, members)
                if prefix is None or prefix not in PREFIX_MAP:
                    unparsed.append((code, scode))
                    continue
                indicator, subcomp, unit = PREFIX_MAP[prefix]
                mname = members[mkey]
                try:
                    val = float(v)
                except ValueError:
                    continue
                dt = datetime.strptime(d, "%Y-%m-%d").date()
                q = f"{dt.year}Q{(dt.month - 1)//3 + 1}"
                rows.append({
                    "dimension": dim,
                    "member": mname,
                    "member_order": order.index(mkey),
                    "indicator": indicator,
                    "subcomponent": subcomp,
                    "unit": unit,
                    "date": pd.Timestamp(dt),
                    "quarter": q,
                    "value": val,
                    "series_code": scode,
                    "series_label": sdd.get("label", ""),
                    "group_code": code,
                    "group_label": gd.get("label", ""),
                    "page_link": page_link,
                    "valet_url": valet_url,
                })
                g_inds.add(indicator)
                g_members.add(mname)
                g_units.add(unit)
                g_dates.append(dt)
                definitions.setdefault(prefix, {
                    "indicator": indicator, "subcomponent": subcomp, "unit": unit,
                    "description": sdd.get("description", ""),
                })

        catalog.append({
            "group_code": code,
            "dimension": dim,
            "group_label": gd.get("label", ""),
            "indicators": ", ".join(sorted(g_inds)),
            "n_series": len(sdet),
            "members": ", ".join(sorted(g_members)),
            "units": ", ".join(sorted(g_units)),
            "date_min": min(g_dates).isoformat() if g_dates else "",
            "date_max": max(g_dates).isoformat() if g_dates else "",
            "n_obs": len(obs),
            "page_link": page_link,
            "valet_url": valet_url,
            "group_description": gdesc,
        })

    tidy = pd.DataFrame(rows)
    cat = pd.DataFrame(catalog).sort_values(["dimension", "group_code"]).reset_index(drop=True)
    return tidy, cat, definitions, unparsed


# ---------------------------------------------------------------------------
# Shape helpers
# ---------------------------------------------------------------------------
def order_tidy(df: pd.DataFrame) -> pd.DataFrame:
    ind_rank = {k: i for i, k in enumerate(INDICATOR_ORDER)}
    df = df.copy()
    df["_ir"] = df["indicator"].map(ind_rank).fillna(999).astype(int)
    df = df.sort_values(["_ir", "subcomponent", "member_order", "date"])
    df = df.drop(columns=["_ir", "member_order"])
    cols = ["dimension", "member", "indicator", "subcomponent", "unit",
            "date", "quarter", "value", "series_code", "series_label",
            "group_code", "group_label", "page_link", "valet_url"]
    return df[cols].reset_index(drop=True)


def make_wide(df: pd.DataFrame, member_order: list[str], member_map: dict) -> pd.DataFrame:
    """Pivot to Date x (indicator, subcomponent, member)."""
    if df.empty:
        return pd.DataFrame()
    ind_rank = {k: i for i, k in enumerate(INDICATOR_ORDER)}
    name_to_rank = {v: member_order.index(k) for k, v in member_map.items()}
    d = df.copy()
    d["col_ind"] = d["indicator"].map(lambda x: f"{ind_rank.get(x,999):02d}|{x}")
    d["col_mem"] = d["member"].map(lambda x: f"{name_to_rank.get(x,99):02d}|{x}")
    wide = d.pivot_table(
        index="date",
        columns=["col_ind", "subcomponent", "col_mem"],
        values="value", aggfunc="first",
    )
    wide = wide.sort_index(axis=1)
    # strip the sort keys from the labels
    wide.columns = pd.MultiIndex.from_tuples(
        [(ci.split("|", 1)[1], sc, cm.split("|", 1)[1]) for ci, sc, cm in wide.columns],
        names=["indicator", "subcomponent", "member"],
    )
    wide = wide.sort_index()
    return wide


# ---------------------------------------------------------------------------
# Notes
# ---------------------------------------------------------------------------
CAVEAT = ("Because the responses are presented here as four-quarter moving averages, "
          "these charts may not show the same movements as the aggregate BOS data "
          "published each quarter.")

def notes_frame(tidy: pd.DataFrame, cat: pd.DataFrame) -> pd.DataFrame:
    dmin = tidy["date"].min().date().isoformat() if not tidy.empty else ""
    dmax = tidy["date"].max().date().isoformat() if not tidy.empty else ""
    lines = [
        ("Title", "Bank of Canada Business Outlook Survey (BOS) — disaggregated series"),
        ("Contents", "BOS results broken down by sector, region and firm size "
                     "(quarterly four-quarter moving averages)."),
        ("Source", "Bank of Canada Valet API (public, no key): "
                   "https://www.bankofcanada.ca/valet"),
        ("Retrieved", date.today().isoformat()),
        ("Groups used", f"{len(cat)} Valet groups "
                        "(BOS_BYSECTOR_*, BOS_BYREGION_*, BOS_BYSIZE_*, BOS_REGIONAL_INDICATOR)"),
        ("Date coverage", f"{dmin} to {dmax} (varies by indicator; some series discontinued earlier)"),
        ("Frequency", "Quarterly (dates are quarter-start: Jan/Apr/Jul/Oct)"),
        ("IMPORTANT — units caveat", CAVEAT),
        ("Units", "Mixed. 'Balance of opinion' = % higher/more minus % lower/less (can be negative). "
                  "'% of firms' = share of respondents. 'Standardized units' = regional indicator scale. "
                  "Unit is carried per row in the tidy sheets; never mix units on one chart axis."),
        ("Sector definitions (NAICS)", "Primary: 11, 21 | Manufacturing: 311-339 | "
                                       "CITU (construction, information, transportation, utilities): 22, 23, 48, 49, 51 | "
                                       "Trade: 41, 44, 45 | FIRE (finance, insurance, real estate): 52, 53 | "
                                       "CPBS (commercial, personal & business services): 54, 55, 56, 71, 72, 81"),
        ("Regions", "Atlantic, Quebec, Ontario, Prairies, British Columbia"),
        ("Firm sizes", "Small, Medium-sized, Large"),
        ("Regional BOS indicator", "BOS_REGIONAL_INDICATOR gives each region's contribution to a "
                                    "standardized regional indicator (different scale/units and shorter "
                                    "history than the by-region balance-of-opinion series)."),
        ("Sheets", "Notes; Catalog (one row per Valet group); Definitions (per-indicator survey "
                   "question); Sector_tidy/Sector_wide; Region_tidy/Region_wide; "
                   "Size_tidy/Size_wide. '_tidy' = long format (one row per series x date); "
                   "'_wide' = Date x series matrix."),
        ("Vintage note", "These are the current BoC-published moving-average panels, not point-in-time "
                         "vintages; values reflect the latest revision/standardization."),
    ]
    return pd.DataFrame(lines, columns=["Field", "Value"])


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Fetching + parsing BOS disaggregated groups ...")
    tidy, cat, definitions, unparsed = parse_all()
    if unparsed:
        print(f"  WARNING: {len(unparsed)} unparsed series (logged):",
              sorted(set(unparsed))[:10])

    sector = order_tidy(tidy[tidy["dimension"] == "Sector"])
    region = order_tidy(tidy[tidy["dimension"] == "Region"])
    size = order_tidy(tidy[tidy["dimension"] == "Firm size"])

    sector_w = make_wide(tidy[tidy["dimension"] == "Sector"], SECTOR_ORDER, SECTOR_MEMBERS)
    # region wide: combine both region member maps for ordering
    reg_map = {**REGION_MEMBERS, **{"ATLANTIC": "Atlantic", "PRAIRIES": "Prairies",
                                    "INDICATOR": "All regions (indicator)"}}
    reg_order = ["AT", "ATLANTIC", "QC", "ON", "PR", "PRAIRIES", "BC", "INDICATOR"]
    region_w = make_wide(tidy[tidy["dimension"] == "Region"], reg_order, reg_map)
    size_w = make_wide(tidy[tidy["dimension"] == "Firm size"], SIZE_ORDER, SIZE_MEMBERS)

    defs_rows = []
    for prefix, info in definitions.items():
        defs_rows.append({
            "indicator": info["indicator"],
            "subcomponent": info["subcomponent"],
            "unit": info["unit"],
            "code_prefix": prefix,
            "survey_definition": info["description"],
        })
    ind_rank = {k: i for i, k in enumerate(INDICATOR_ORDER)}
    defs = (pd.DataFrame(defs_rows)
            .assign(_r=lambda d: d["indicator"].map(lambda x: ind_rank.get(x, 999)))
            .sort_values(["_r", "subcomponent"]).drop(columns="_r").reset_index(drop=True))

    notes = notes_frame(tidy, cat)

    print(f"Writing {OUT_XLSX} ...")
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl", datetime_format="yyyy-mm-dd") as xl:
        notes.to_excel(xl, sheet_name="Notes", index=False)
        cat.to_excel(xl, sheet_name="Catalog", index=False)
        defs.to_excel(xl, sheet_name="Definitions", index=False)
        sector.to_excel(xl, sheet_name="Sector_tidy", index=False)
        sector_w.to_excel(xl, sheet_name="Sector_wide")
        region.to_excel(xl, sheet_name="Region_tidy", index=False)
        region_w.to_excel(xl, sheet_name="Region_wide")
        size.to_excel(xl, sheet_name="Size_tidy", index=False)
        size_w.to_excel(xl, sheet_name="Size_wide")
        _autofit(xl)

    # ---- verification summary ----
    print("\n=== VERIFICATION ===")
    for name, df in [("Sector", sector), ("Region", region), ("Firm size", size)]:
        if df.empty:
            print(f"{name}: EMPTY"); continue
        print(f"{name:9s}: rows={len(df):5d}  series={df['series_code'].nunique():3d}  "
              f"indicators={df['indicator'].nunique():2d}  members={df['member'].nunique()}  "
              f"dates {df['date'].min().date()}..{df['date'].max().date()}")
    print(f"Catalog groups: {len(cat)}   Definitions: {len(defs)}")
    print("Total tidy rows:", len(tidy))
    return tidy, cat


def _autofit(xl):
    from openpyxl.utils import get_column_letter
    for ws in xl.book.worksheets:
        max_row = min(ws.max_row, 200)
        for ci in range(1, ws.max_column + 1):
            length = 0
            for ri in range(1, max_row + 1):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    length = max(length, len(str(v)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max(length + 2, 10), 60)


if __name__ == "__main__":
    main()
